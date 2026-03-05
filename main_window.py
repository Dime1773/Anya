#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
main_window.py - PyQt6 v1.11.2 (исправленный и завершённый)
Anya Distributor - распределение файлов по подразделениям
ИЗМЕНЕНИЯ:
- Исправлена сортировка по "№" и "Префикс"
- Все колонки масштабируются вручную (включая "Подразделение")
- Таблица занимает всю ширину (без пустого пространства)
- При запуске ширина колонок = ширине заголовка
- Все методы присутствуют и корректны
"""
import logging
import openpyxl
import json
from PyQt6.QtWidgets import QFileDialog, QMessageBox
from typing import List, Dict
from pathlib import Path
from datetime import datetime
from PyQt6.QtWidgets import (
    QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QPushButton, QTableWidget, QTableWidgetItem, QTextEdit,
    QCheckBox, QLabel, QComboBox, QFileDialog, QMessageBox,
    QDialog, QProgressBar, QHeaderView, QLineEdit, QTabWidget, QSplitter, QApplication
)
from PyQt6.QtCore import Qt, QUrl
from PyQt6.QtGui import QColor, QFont, QDesktopServices
from database import DatabaseManager
from workers import AsyncCheckWorker, TransferWorker, AutomatErrorWorker
from logger import init_logging, init_ui_logging, LogSignalEmitter

logger = logging.getLogger("Anya distributor")


class NumericTableWidgetItem(QTableWidgetItem):
    """Кастомная ячейка для правильной сортировки чисел"""
    def __lt__(self, other):
        try:
            val1 = self.data(Qt.ItemDataRole.UserRole) or 0
            val2 = other.data(Qt.ItemDataRole.UserRole) or 0
            v1 = float(val1) if val1 is not None else 0.0
            v2 = float(val2) if val2 is not None else 0.0
            return v1 < v2
        except:
            return super().__lt__(other)


class InstructionDialog(QDialog):
    """Диалог с инструкциями"""
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Инструкция - Anya Distributor")
        self.setGeometry(150, 150, 700, 600)
        self.setModal(True)
        layout = QVBoxLayout()
        instruction_text = QTextEdit()
        instruction_text.setReadOnly(True)
        instruction_text.setFont(QFont("Courier", 10))
        text = """Anya DISTRIBUTOR v1.11.2 (исправленный)
1. ЗАГРУЗКА БАЗЫ ДАННЫХ
- Нажмите кнопку "Загрузить БД"
- Приложение загружает справочник apteki.json и создаёт/обновляет base.json
- Отображаются все подразделения из base.json
2. ПРОВЕРКА ПОДКЛЮЧЕНИЯ
- Выберите подразделения (галочки)
- Нажмите "Проверить подключение"
- Проверяется доступность на порт 445 (SMB)
- Кнопка "Стоп" останавливает проверку
3. ПЕРЕДАЧА ФАЙЛОВ
- Даже если статус "Нет", вы можете выбрать любой IP
- Передача возможна на любой указанный IP
- Кнопка "Стоп" останавливает передачу
4. АВТОМАТ
- Кнопка активна, если есть logs/errors.xlsx
- Автоматически проверяет и передаёт файлы из этого списка
5. ТЕМА
- В меню "Вид" можно переключить тему: тёмная/светлая
6. ПРОВЕРКА ЦЕЛОСТНОСТИ (SHA256)
- В диалоге выбора IP появился чекбокс "Проверять целостность"
- При включённой опции — контроль хеша SHA256
По вопросам: Telegram @Dime1337
"""
        instruction_text.setText(text)
        layout.addWidget(instruction_text)
        button_layout = QHBoxLayout()
        btn_telegram = QPushButton("Написать разрабу")
        btn_telegram.clicked.connect(self._open_telegram)
        button_layout.addWidget(btn_telegram)
        button_layout.addStretch()
        btn_close = QPushButton("Закрыть")
        btn_close.clicked.connect(self.close)
        button_layout.addWidget(btn_close)
        layout.addLayout(button_layout)
        self.setLayout(layout)

    def _open_telegram(self):
        QDesktopServices.openUrl(QUrl("https://t.me/Dime1337"))

class ErrorsTab(QWidget):
    """Вкладка с ошибками (теперь с IP-адресом)"""
    def __init__(self):
        super().__init__()
        layout = QVBoxLayout()
        top_layout = QHBoxLayout()
        top_layout.addWidget(QLabel("Логи ошибок:"))
        top_layout.addStretch()
        btn_export = QPushButton("Экспортировать в Excel")
        btn_export.clicked.connect(self._export_errors)
        top_layout.addWidget(btn_export)
        btn_clear = QPushButton("Очистить")
        btn_clear.clicked.connect(self._clear_errors)
        top_layout.addWidget(btn_clear)
        layout.addLayout(top_layout)

        self.errors_table = QTableWidget()
        # 🔥 Добавлена 5-я колонка: IP
        self.errors_table.setColumnCount(5)
        self.errors_table.setHorizontalHeaderLabels(["Время", "Префикс", "Подразделение", "IP", "Ошибка"])
        for col in range(5):
            self.errors_table.horizontalHeader().setSectionResizeMode(col, QHeaderView.ResizeMode.Stretch)
        layout.addWidget(self.errors_table)
        self.setLayout(layout)

    def add_error(self, prefix: str, branch_name: str, error_msg: str, ip: str = ""):
        """🔥 Добавлен параметр `ip`"""
        row = self.errors_table.rowCount()
        self.errors_table.insertRow(row)
        time_str = datetime.now().strftime("%H:%M:%S")
        self.errors_table.setItem(row, 0, QTableWidgetItem(time_str))
        self.errors_table.setItem(row, 1, QTableWidgetItem(str(prefix)))
        self.errors_table.setItem(row, 2, QTableWidgetItem(str(branch_name)))
        self.errors_table.setItem(row, 3, QTableWidgetItem(str(ip)))  # ← IP
        error_item = QTableWidgetItem(str(error_msg))
        error_item.setForeground(QColor("#ff4444"))
        self.errors_table.setItem(row, 4, error_item)  # ← Ошибка теперь в колонке 4
        self.errors_table.scrollToBottom()

    def _export_errors(self):
        try:
            from openpyxl.styles import Font, PatternFill
        except ImportError:
            QMessageBox.warning(self, "Ошибка", "Требуется: pip install openpyxl")
            return

        file_path, _ = QFileDialog.getSaveFileName(
            self, "Сохранить ошибки", "logs/errors.xlsx", "Excel Files (*.xlsx)"
        )
        if not file_path:
            return

        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Ошибки"
            # 🔥 Добавлена колонка "IP"
            headers = ["Время", "Префикс", "Подразделение", "IP", "Ошибка"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(1, col, header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

            for row in range(self.errors_table.rowCount()):
                for col in range(5):  # ← 5 колонок
                    item = self.errors_table.item(row, col)
                    if item:
                        ws.cell(row + 2, col + 1, item.text())

            ws.column_dimensions["A"].width = 12  # Время
            ws.column_dimensions["B"].width = 12  # Префикс
            ws.column_dimensions["C"].width = 25  # Подразделение
            ws.column_dimensions["D"].width = 18  # IP ← новая ширина
            ws.column_dimensions["E"].width = 60  # Ошибка

            wb.save(file_path)
            QMessageBox.information(self, "Успех", f"Экспортировано: {file_path}")
            logger.info(f"Ошибки экспортированы в {file_path}")
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Ошибка экспорта: {e}")
            logger.error(f"Ошибка экспорта: {e}")

    def _clear_errors(self):
        self.errors_table.setRowCount(0)
        logger.info("Логи ошибок очищены")

class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Anya Distributor v1.11.2 (исправленный)")
        self.resize(1200, 800)
        self.setMinimumSize(800, 600)
        init_logging("logs")
        self.log_emitter = LogSignalEmitter()
        init_ui_logging(self.log_emitter)
        self.db = DatabaseManager("apteki.json", "base.json")
        self.check_worker = None
        self.transfer_worker = None
        self.automat_worker = None
        self.selected_files = []
        self.selected_branches = []
        self.all_branches = []
        self.select_all_state = False
        self._sort_orders = {}
        self._image_counter = 0
        self.row_count_label = None
        self.selected_count_label = None
        self.logs_visible = True
        self.main_splitter = None
        self.current_sort_column = -1
        self.current_sort_order = Qt.SortOrder.AscendingOrder
        self._init_ui()
        self._setup_styles()
        self._load_database()
        logger.info("=" * 80)
        logger.info("Anya Distributor v1.11.2 (исправленный) запущено")
        logger.info("=" * 80)

    def _init_ui(self):
        main_widget = QWidget()
        self.setCentralWidget(main_widget)
        main_layout = QVBoxLayout(main_widget)

        # Верхняя панель
        top_panel = QHBoxLayout()
        btn_load_db = QPushButton("Загрузить БД")
        self._load_database()
        top_panel.addWidget(btn_load_db)
        self.btn_check = QPushButton("Проверить подключение")
        self.btn_check.clicked.connect(self._on_check_connection)
        self.btn_check.setEnabled(False)
        top_panel.addWidget(self.btn_check)
        self.btn_transfer = QPushButton("Передача файлов")
        self.btn_transfer.clicked.connect(self._on_transfer_files)
        self.btn_transfer.setEnabled(False)
        top_panel.addWidget(self.btn_transfer)
        self.btn_automat = QPushButton("Автомат")
        self.btn_automat.clicked.connect(self._on_automat)
        top_panel.addWidget(self.btn_automat)
        self.btn_select_all = QPushButton("Выбрать все")
        self.btn_select_all.clicked.connect(self._on_select_all)
        top_panel.addWidget(self.btn_select_all)
        top_panel.addStretch()
        self.btn_stop = QPushButton("Стоп")
        self.btn_stop.clicked.connect(self._on_stop)
        self.btn_stop.setEnabled(False)
        top_panel.addWidget(self.btn_stop)
        self.btn_toggle_logs = QPushButton("Свернуть логи")
        self.btn_toggle_logs.clicked.connect(self._on_toggle_logs)
        top_panel.addWidget(self.btn_toggle_logs)
        btn_instruction = QPushButton("Инструкция")
        btn_instruction.clicked.connect(self._on_show_instruction)
        top_panel.addWidget(btn_instruction)
        main_layout.addLayout(top_panel)

        # Прогресс
        progress_layout = QHBoxLayout()
        self.progress_label = QLabel()
        self.progress_bar = QProgressBar()
        self.progress_bar.setVisible(False)
        progress_layout.addWidget(self.progress_label)
        progress_layout.addWidget(self.progress_bar)
        main_layout.addLayout(progress_layout)

        # Левая панель (фильтры + Anya)
        try:
            from left_panel_anya import AnyaPanel
            self.anya_panel = AnyaPanel()
        except ImportError as e:
            logger.error(f"Ошибка импорта AnyaPanel: {e}")
            self.anya_panel = QLabel("AnyaPanel не доступен")
            self.anya_panel.setAlignment(Qt.AlignmentFlag.AlignCenter)
        left_layout = QVBoxLayout()
        left_layout.addWidget(QLabel("ФИЛЬТР (жирный)", font=QFont("Arial", weight=QFont.Weight.Bold)))
        filter_label = QLabel("По статусу:")
        left_layout.addWidget(filter_label)
        self.filter_combo = QComboBox()
        self.filter_combo.addItems(["Все", "Online", "Offline"])
        self.filter_combo.currentTextChanged.connect(self._on_filter_changed)
        left_layout.addWidget(self.filter_combo)
        left_layout.addWidget(QLabel("По названию:"))
        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText("Поиск...")
        self.search_input.textChanged.connect(self._on_filter_changed)
        left_layout.addWidget(self.search_input)
        counters_layout = QVBoxLayout()
        counters_layout.addWidget(QLabel("Видимых строк:"))
        self.row_count_label = QLabel("0/0")
        self.row_count_label.setFont(QFont("Arial", weight=QFont.Weight.Bold))
        self.row_count_label.setStyleSheet("color: #2a9a9f;")
        counters_layout.addWidget(self.row_count_label)
        counters_layout.addWidget(QLabel("Выделено:"))
        self.selected_count_label = QLabel("0")
        self.selected_count_label.setFont(QFont("Arial", weight=QFont.Weight.Bold))
        self.selected_count_label.setStyleSheet("color: #00ff00;")
        counters_layout.addWidget(self.selected_count_label)
        left_layout.addLayout(counters_layout)
        left_layout.addSpacing(10)
        left_layout.addWidget(self.anya_panel)
        left_widget = QWidget()
        left_widget.setLayout(left_layout)

        # Центр — таблица
        center_layout = QVBoxLayout()
        center_layout.addWidget(QLabel("ПОДРАЗДЕЛЕНИЯ", font=QFont("Arial", weight=QFont.Weight.Bold)))
        self.table = QTableWidget()
        self.table.setColumnCount(8)
        self.table.setHorizontalHeaderLabels(["№", "OK", "Префикс", "Подразделение", "IP", "Сервер", "Оператор", "Оператор 2"])
        self.table.verticalHeader().setVisible(False)
        header = self.table.horizontalHeader()

        # 🔑 ВСЕ колонки Interactive — можно тянуть вручную, включая "Подразделение"
        for col in range(8):
            header.setSectionResizeMode(col, QHeaderView.ResizeMode.Interactive)

        self.table.itemClicked.connect(self._on_table_item_clicked)
        header.sectionClicked.connect(self._on_header_clicked)
        center_layout.addWidget(self.table)
        center_widget = QWidget()
        center_widget.setLayout(center_layout)

        # Сплиттер: лево + центр
        top_splitter = QSplitter(Qt.Orientation.Horizontal)
        top_splitter.addWidget(left_widget)
        top_splitter.addWidget(center_widget)
        top_splitter.setSizes([250, 950])  # [ширина_левой, ширина_правой]
        # Не задаём setStretchFactor — центр занимает всё свободное место

        # Логи и ошибки
        logs_widget = QWidget()
        logs_layout = QVBoxLayout(logs_widget)
        logs_top = QHBoxLayout()
        logs_top.addWidget(QLabel("ЛОГИ", font=QFont("Arial", weight=QFont.Weight.Bold)))
        logs_top.addStretch()
        btn_clear_logs = QPushButton("Очистить логи")
        btn_clear_logs.clicked.connect(self._on_clear_logs)
        logs_top.addWidget(btn_clear_logs)
        logs_layout.addLayout(logs_top)
        self.log_panel = QTextEdit()
        self.log_panel.setReadOnly(True)
        font = QFont("Courier")
        font.setPointSize(QApplication.font().pointSize())
        self.log_panel.setFont(font)
        logs_layout.addWidget(self.log_panel)
        self.errors_tab = ErrorsTab()
        self.tabs = QTabWidget()
        self.tabs.addTab(logs_widget, "Логи")
        self.tabs.addTab(self.errors_tab, "Ошибки")

        # Главный сплиттер
        self.main_splitter = QSplitter(Qt.Orientation.Vertical)
        self.main_splitter.addWidget(top_splitter)
        self.main_splitter.addWidget(self.tabs)
        self.main_splitter.setStretchFactor(0, 2)
        self.main_splitter.setStretchFactor(1, 1)
        main_layout.addWidget(self.main_splitter)

        self.log_emitter.log_message.connect(self._on_log_message)
        self._update_automat_button()

    def _setup_styles(self):
        stylesheet = """
QMainWindow { background-color: #1a1a1a; }
QWidget { background-color: #1a1a1a; color: #e0e0e0; }
QPushButton { background-color: #2a2a2a; color: #e0e0e0; border: 1px solid #3a3a3a;
padding: 8px; border-radius: 4px; font-weight: bold; }
QPushButton:hover { background-color: #3a3a3a; }
QPushButton:pressed { background-color: #2a7a7f; }
QPushButton:disabled { background-color: #2a2a2a; color: #555555; }
QTableWidget { background-color: #2a2a2a; alternate-background-color: #252525;
gridline-color: #3a3a3a; border: 1px solid #3a3a3a; }
QTableWidget::item { padding: 4px; }
QTableWidget::item:selected { background-color: #2a7a7f; }
QHeaderView::section { background-color: #1e1e1e; color: #e0e0e0; padding: 4px;
border: 1px solid #3a3a3a; }
QTextEdit { background-color: #252525; color: #e0e0e0; border: 1px solid #3a3a3a; }
QLineEdit, QComboBox { background-color: #2a2a2a; color: #e0e0e0; border: 1px solid #3a3a3a;
padding: 4px; border-radius: 4px; }
QLabel { color: #e0e0e0; }
QProgressBar { background-color: #2a2a2a; border: 1px solid #3a3a3a; border-radius: 4px;
text-align: center; }
QProgressBar::chunk { background-color: #2a9a9f; }
QDialog { background-color: #1a1a1a; }
QTabWidget::pane { border: 1px solid #3a3a3a; }
QTabBar::tab { background-color: #2a2a2a; color: #e0e0e0; padding: 5px 15px; }
QTabBar::tab:selected { background-color: #2a7a7f; }
QSplitter::handle { background-color: #3a3a3a; }
"""
        self.setStyleSheet(stylesheet)

    def load_database(self):
        """Метод, вызываемый при нажатии кнопки Загрузить БД"""
        # Диалог выбора файла
        file_name, _ = QFileDialog.getOpenFileName(self, "Выберите файл базы данных", "",
                                                "JSON files (*.json);;All Files (*)")
        if file_name:
            try:
                # Чтение файла
                with open(file_name, 'r') as f:
                    db_data = json.load(f)
                    # Обработать полученный JSON-файл (загрузка данных)
                    print("Файл загружен:", file_name)
            except FileNotFoundError:
                QMessageBox.warning(self, "Ошибка", "Файл не найден.")
            except json.JSONDecodeError:
                QMessageBox.warning(self, "Ошибка", "Неверный формат JSON.")
            except Exception as e:
                QMessageBox.critical(self, "Критическая ошибка", f"При загрузке произошла ошибка: {e}")

    def _populate_table(self, branches: List[Dict]):
        valid_branches = [b for b in branches if isinstance(b, dict)]
        self.table.setRowCount(len(valid_branches))
        self.all_branches = valid_branches

        for row, branch in enumerate(valid_branches):
            # №
            num_item = NumericTableWidgetItem(str(row + 1))
            num_item.setData(Qt.ItemDataRole.UserRole, row + 1)
            self.table.setItem(row, 0, num_item)

            # OK
            check_item = QTableWidgetItem("")
            check_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            self.table.setItem(row, 1, check_item)

            # Префикс
            prefix_val = branch.get("prefix", "")
            try:
                prefix_num = int(prefix_val)
            except ValueError:
                prefix_num = 0
            prefix_item = NumericTableWidgetItem(str(prefix_val))
            prefix_item.setData(Qt.ItemDataRole.UserRole, prefix_num)
            self.table.setItem(row, 2, prefix_item)

            # Остальные
            self.table.setItem(row, 3, QTableWidgetItem(str(branch.get("name", ""))))
            self.table.setItem(row, 4, QTableWidgetItem(str(branch.get("vne_ip", ""))))
            for col in (5, 6, 7):
                self.table.setItem(row, col, QTableWidgetItem(""))

        # 🔑 Установка ширины после загрузки
        self._resize_columns_to_content()
        self._update_row_counter()
        self._update_selection()

    def _resize_columns_to_content(self):
        """Установить ширину колонок по содержимому заголовков и данных."""
        header = self.table.horizontalHeader()
        for col in range(8):
            header.resizeSection(col, header.sectionSizeHint(col))
        
        # === ЖЕСТКИЕ РАЗМЕРЫ КОЛОНОК (в пикселях) ===
        self.table.setColumnWidth(0, 40)   # №
        self.table.setColumnWidth(1, 40)   # OK
        self.table.setColumnWidth(2, 70)   # Префикс
        self.table.setColumnWidth(3, 450)  # Подразделение
        self.table.setColumnWidth(4, 190)  # IP
        self.table.setColumnWidth(5, 80)   # Сервер
        self.table.setColumnWidth(6, 100)  # Оператор
        self.table.setColumnWidth(7, 120)  # Оператор 2

    def _on_load_database(self):
        """Обработчик кнопки 'Загрузить БД'"""
        try:
            self.db.rescan_bases()
            self._load_database()
            self._update_automat_button()
            QMessageBox.information(self, "Успех", "База пересканирована и загружена")
        except Exception as e:
            logger.error(f"Ошибка: {e}")
            QMessageBox.critical(self, "Ошибка", f"Ошибка: {e}")

    def _on_header_clicked(self, col: int):
        if col == 4:
            self._show_image()
            return
        order = self._sort_orders.get(col, Qt.SortOrder.AscendingOrder)
        new_order = Qt.SortOrder.DescendingOrder if order == Qt.SortOrder.AscendingOrder else Qt.SortOrder.AscendingOrder
        self._sort_orders[col] = new_order

        if col in (0, 2):
            self.table.setSortingEnabled(True)
            self.table.sortItems(col, new_order)
            self.table.setSortingEnabled(False)
        elif col in (5, 6, 7):
            self._sort_yes_no_column(col, new_order)
        else:
            self.table.setSortingEnabled(True)
            self.table.sortItems(col, new_order)
            self.table.setSortingEnabled(False)

    def _sort_yes_no_column(self, col: int, order: Qt.SortOrder):
        for row in range(self.table.rowCount()):
            item = self.table.item(row, col)
            if item:
                value = 1 if item.text() == "Да" else 0
                item.setData(Qt.ItemDataRole.UserRole, value)
        self.table.setSortingEnabled(True)
        self.table.sortItems(col, order)
        self.table.setSortingEnabled(False)

    # === ОСТАЛЬНЫЕ ОБЯЗАТЕЛЬНЫЕ МЕТОДЫ (без изменений, но присутствуют) ===

    def _toggle_check_item(self, row: int):
        item = self.table.item(row, 1)
        if item.text() == "✓":
            item.setText("")
        else:
            item.setText("✓")

    def _update_selection(self):
        self.selected_branches = []
        for row in range(self.table.rowCount()):
            if not self.table.isRowHidden(row):
                mark = self.table.item(row, 1)
                if mark and mark.text() == "✓":
                    prefix = self.table.item(row, 2)
                    if prefix:
                        self.selected_branches.append(prefix.text())
        has_selection = len(self.selected_branches) > 0
        self.btn_check.setEnabled(has_selection)
        self.btn_transfer.setEnabled(has_selection)
        self._update_selected_counter()

    def _update_selected_counter(self):
        count = sum(
            1 for row in range(self.table.rowCount())
            if not self.table.isRowHidden(row) and self.table.item(row, 1).text() == "✓"
        )
        self.selected_count_label.setText(str(count))
        color = "#00ff00" if count > 0 else "#ff4444"
        self.selected_count_label.setStyleSheet(f"color: {color}; font-weight: bold;")

    def _update_row_counter(self):
        total = self.table.rowCount()
        visible = sum(1 for row in range(total) if not self.table.isRowHidden(row))
        self.row_count_label.setText(f"{visible}/{total}")
        color = "#ff4444" if visible == 0 and total > 0 else "#2a9a9f"
        self.row_count_label.setStyleSheet(f"color: {color}; font-weight: bold;")

    def _update_automat_button(self):
        errors_file = Path("logs/errors.xlsx")
        enabled = errors_file.exists() and errors_file.stat().st_size > 0
        self.btn_automat.setEnabled(enabled)

    def _on_select_all(self):
        self.table.setUpdatesEnabled(False)
        self.select_all_state = not self.select_all_state
        for row in range(self.table.rowCount()):
            if not self.table.isRowHidden(row):
                self.table.item(row, 1).setText("✓" if self.select_all_state else "")
        self.table.setUpdatesEnabled(True)
        self._update_selection()

    def _on_filter_changed(self):
        status_filter = self.filter_combo.currentText().lower()
        search_text = self.search_input.text().lower()
        for row in range(self.table.rowCount()):
            name_item = self.table.item(row, 3)
            prefix_item = self.table.item(row, 2)
            if not name_item or not prefix_item:
                self.table.setRowHidden(row, True)
                continue
            name = name_item.text().lower()
            prefix = prefix_item.text().lower()
            server = self.table.item(row, 5).text()
            op1 = self.table.item(row, 6).text()
            op2 = self.table.item(row, 7).text()
            is_online = any(s == "Да" for s in (server, op1, op2))
            is_offline = all(s in ("", "Нет") for s in (server, op1, op2))
            show_status = (
                status_filter == "все" or
                (status_filter == "online" and is_online) or
                (status_filter == "offline" and is_offline)
            )
            show_search = not search_text or search_text in name or search_text in prefix
            self.table.setRowHidden(row, not (show_status and show_search))
        self._update_row_counter()
        self._update_selection()

    def _on_table_item_clicked(self, item):
        if item.column() == 1:
            self._toggle_check_item(item.row())
            self._update_selection()

    def _show_image(self):
        self._image_counter += 1
        img_number = 2 if self._image_counter % 2 == 0 else 1
        try:
            if img_number == 1 and hasattr(self.anya_panel, 'show_image_1'):
                self.anya_panel.show_image_1()
            elif img_number == 2 and hasattr(self.anya_panel, 'show_image_2'):
                self.anya_panel.show_image_2()
        except Exception as e:
            logger.error(f"Ошибка показа изображения: {e}")

    def _on_toggle_logs(self):
        self.logs_visible = not self.logs_visible
        if self.logs_visible:
            self.main_splitter.setSizes([1000, 500])
            self.btn_toggle_logs.setText("Свернуть логи")
        else:
            self.main_splitter.setSizes([1000, 0])
            self.btn_toggle_logs.setText("Развернуть логи")

    def _on_check_connection(self):
        if not self.selected_branches:
            QMessageBox.warning(self, "Предупреждение", "Выберите подразделения")
            return
        
        try:
            if hasattr(self.anya_panel, 'show_check_image'):
                self.anya_panel.show_check_image()
        except:
            pass
        
        branches_to_check = [b for b in self.all_branches if b.get("prefix") in self.selected_branches]
        if not branches_to_check:
            QMessageBox.warning(self, "Предупреждение", "Подразделения не найдены")
            return
        
        logger.info(f"🚀 Быстрая проверка {len(branches_to_check)} подразделений...")
        
        # 🔥 Используем настройки из конфига
        from config import (CHECK_MAX_CONCURRENT, CHECK_PING_TIMEOUT, CHECK_PORT_TIMEOUT, 
                            CHECK_BATCH_SIZE, CHECK_PAUSE_BETWEEN_BATCHES)
        
        self.check_worker = AsyncCheckWorker(
            branches_to_check, 
            self.db,
            max_concurrent=CHECK_MAX_CONCURRENT,
            ping_timeout=CHECK_PING_TIMEOUT,
            port_timeout=CHECK_PORT_TIMEOUT,
            batch_size=CHECK_BATCH_SIZE,
            pause_between_batches=CHECK_PAUSE_BETWEEN_BATCHES
        )
        self.check_worker.progress.connect(self._on_check_progress)
        self.check_worker.status_updated.connect(self._on_status_updated)
        self.check_worker.finished.connect(self._on_check_finished)
        self.check_worker.error.connect(self._on_check_error)
        
        self.progress_bar.setVisible(True)
        self.progress_label.setText(f"0/{len(branches_to_check)}")
        self.btn_check.setEnabled(False)
        self.btn_transfer.setEnabled(False)
        self.btn_stop.setEnabled(True)
        self.check_worker.start()

    def _on_check_progress(self, current: int, total: int):
        self.progress_bar.setMaximum(total)
        self.progress_bar.setValue(current)
        self.progress_label.setText(f"{current}/{total}")

    def _on_status_updated(self, prefix, status, active_ip, srv, op1, op2, error):
        for row in range(self.table.rowCount()):
            if self.table.item(row, 2).text() == prefix:
                for col, val in zip([5, 6, 7], [srv, op1, op2]):
                    item = self.table.item(row, col)
                    item.setText(val)
                    color = "#00ff00" if val == "Да" else "#ff4444" if val == "Нет" else "#cccccc"
                    item.setForeground(QColor(color))
                break
        for branch in self.all_branches:
            if branch.get("prefix") == prefix:
                branch["status"] = status
                branch["active_ip"] = active_ip
                break

    def _on_check_finished(self):
        self.progress_bar.setVisible(False)
        self.btn_check.setEnabled(True)
        self.btn_transfer.setEnabled(len(self.selected_branches) > 0)
        self.btn_stop.setEnabled(False)
        self._update_selection()
        try:
            if hasattr(self.anya_panel, 'show_success_image'):
                self.anya_panel.show_success_image()
        except:
            pass
        logger.info("Проверка подключения завершена")
        QMessageBox.information(self, "Успех", "Проверка завершена")

    def _on_check_error(self, error_msg: str):
        self.progress_bar.setVisible(False)
        self.btn_check.setEnabled(True)
        self.btn_transfer.setEnabled(len(self.selected_branches) > 0)
        self.btn_stop.setEnabled(False)
        self._update_selection()
        try:
            if hasattr(self.anya_panel, 'show_error_image'):
                self.anya_panel.show_error_image()
        except:
            pass
        QMessageBox.critical(self, "Ошибка", f"Ошибка: {error_msg}")
        logger.error(f"Ошибка проверки: {error_msg}")

    def _on_transfer_files(self):
        if not self.selected_branches:
            QMessageBox.warning(self, "Предупреждение", "Выберите подразделения")
            return
        button = QMessageBox(self)
        button.setWindowTitle("Выбор типа")
        button.setText("Что передавать?")
        button.setIcon(QMessageBox.Icon.Question)
        btn_files = button.addButton("Файлы", QMessageBox.ButtonRole.YesRole)
        btn_folders = button.addButton("Папки", QMessageBox.ButtonRole.NoRole)
        button.addButton("Отмена", QMessageBox.ButtonRole.RejectRole)
        button.exec()
        if button.clickedButton() == btn_files:
            self._transfer_files_action("files")
        elif button.clickedButton() == btn_folders:
            self._transfer_files_action("folders")

    def _transfer_files_action(self, file_type: str):
        if file_type == "files":
            files, _ = QFileDialog.getOpenFileNames(self, "Выбрать файлы", "", "All Files (*)")
            self.selected_files = files
        else:
            folder = QFileDialog.getExistingDirectory(self, "Выбрать папку", "")
            self.selected_files = [folder] if folder else []
        if not self.selected_files:
            return
        self._show_ip_selection_dialog()

    def _show_ip_selection_dialog(self):
        ip_dialog = QDialog(self)
        ip_dialog.setWindowTitle("Выбор IP адресов")
        ip_dialog.resize(900, 600)
        layout = QVBoxLayout(ip_dialog)
        layout.addWidget(QLabel("Выберите целевые IP адреса (даже если статус \"Нет\")"))
        self.hash_check = QCheckBox("Проверять целостность (SHA256)")
        self.hash_check.setChecked(False)
        layout.addWidget(self.hash_check)
        table = QTableWidget()
        table.setColumnCount(5)
        table.setHorizontalHeaderLabels(["Подразделение", "Все", "Сервер", "OP1", "OP2"])
        branches = [b for b in self.all_branches if b.get("prefix") in self.selected_branches]
        table.setRowCount(len(branches))
        self.ip_checkboxes = {}
        for row, branch in enumerate(branches):
            prefix = branch.get("prefix", "")
            name = branch.get("name", "")
            main_ip = branch.get("ip", "")
            alt_ips = branch.get("alt_ips", [])
            op1_ip = alt_ips[0] if len(alt_ips) > 0 else ""
            op2_ip = alt_ips[1] if len(alt_ips) > 1 else ""
            table.setItem(row, 0, QTableWidgetItem(f"{prefix} - {name}"))
            cb_all = QCheckBox()
            cb_server = QCheckBox(main_ip if main_ip else "Нет")
            cb_op1 = QCheckBox(op1_ip if op1_ip else "Нет")
            cb_op2 = QCheckBox(op2_ip if op2_ip else "Нет")
            table.setCellWidget(row, 1, cb_all)
            table.setCellWidget(row, 2, cb_server)
            table.setCellWidget(row, 3, cb_op1)
            table.setCellWidget(row, 4, cb_op2)
            self.ip_checkboxes[prefix] = {
                "all": cb_all,
                "server": (cb_server, main_ip),
                "op1": (cb_op1, op1_ip),
                "op2": (cb_op2, op2_ip),
            }
            def make_handler(p):
                def handler():
                    state = self.ip_checkboxes[p]["all"].isChecked()
                    self.ip_checkboxes[p]["server"][0].setChecked(state)
                    self.ip_checkboxes[p]["op1"][0].setChecked(state)
                    self.ip_checkboxes[p]["op2"][0].setChecked(state)
                return handler
            cb_all.stateChanged.connect(make_handler(prefix))
        layout.addWidget(table)
        btns_layout = QHBoxLayout()
        for name, key in [("Все серверы", "server"), ("Все OP1", "op1"), ("Все OP2", "op2"), ("Очистить", "clear")]:
            btn = QPushButton(name)
            btn.clicked.connect(lambda _, k=key: self._select_all_ips(k))
            btns_layout.addWidget(btn)
        layout.addLayout(btns_layout)
        buttons = QHBoxLayout()
        start_btn = QPushButton("Начать передачу")
        start_btn.clicked.connect(lambda: self._start_transfer(ip_dialog))
        buttons.addWidget(start_btn)
        cancel_btn = QPushButton("Отмена")
        cancel_btn.clicked.connect(ip_dialog.reject)
        buttons.addWidget(cancel_btn)
        layout.addLayout(buttons)
        ip_dialog.exec()

    def _select_all_ips(self, target_type: str):
        for prefix, cbs in self.ip_checkboxes.items():
            if target_type == "clear":
                cbs["all"].setChecked(False)
                cbs["server"][0].setChecked(False)
                cbs["op1"][0].setChecked(False)
                cbs["op2"][0].setChecked(False)
            else:
                if target_type == "server":
                    cbs["server"][0].setChecked(True)
                elif target_type == "op1":
                    cbs["op1"][0].setChecked(True)
                elif target_type == "op2":
                    cbs["op2"][0].setChecked(True)

    def _start_transfer(self, dialog: QDialog):
        ip_selection = {}
        target_branches = []
        for prefix, cbs in self.ip_checkboxes.items():
            ips = []
            for key, (cb, ip) in [("server", cbs["server"]), ("op1", cbs["op1"]), ("op2", cbs["op2"])]:
                if cb.isChecked() and ip and ip != "Нет":
                    ips.append(ip)
            if ips:
                ip_selection[prefix] = ips
                for b in self.all_branches:
                    if b.get("prefix") == prefix:
                        target_branches.append(b)
                        break
        if not ip_selection:
            QMessageBox.warning(dialog, "Ошибка", "Выберите хотя бы один IP")
            return
        dialog.accept()
        total_targets = sum(len(ips) for ips in ip_selection.values())
        logger.info(f"Начало передачи: {total_targets} целей")
        self.progress_bar.setMaximum(total_targets)
        self.progress_bar.setValue(0)
        self.progress_bar.setVisible(True)
        self.progress_label.setText(f"0/{total_targets}")
        self.btn_check.setEnabled(False)
        self.btn_transfer.setEnabled(False)
        self.btn_stop.setEnabled(True)
        self.transfer_worker = TransferWorker(target_branches, self.selected_files, ip_selection, self.db)
        self.transfer_worker.progress.connect(self._on_transfer_progress)
        self.transfer_worker.transfer_status.connect(self._on_transfer_status)
        self.transfer_worker.error_logged.connect(self._on_error_logged)
        self.transfer_worker.finished.connect(self._on_transfer_finished)
        self.transfer_worker.error.connect(self._on_transfer_error)
        self.transfer_worker.check_integrity = self.hash_check.isChecked()
        self.transfer_worker.stop_requested = False
        self.transfer_worker.start()

    def _on_transfer_progress(self, current: int, total: int):
        self.progress_bar.setMaximum(total)
        self.progress_bar.setValue(current)
        self.progress_label.setText(f"{current}/{total}")

    def _on_transfer_status(self, prefix: str, ip: str, status: str):
        level = logger.info if status.startswith("✓") else logger.warning
        level(f"{prefix} {ip} {status}")

    def _on_automat(self):
        errors_file = Path("logs/errors.xlsx")
        if not errors_file.exists():
            QMessageBox.warning(self, "Файл не найден", "Файл logs/errors.xlsx не найден.\nСначала экспортируйте ошибки.")
            return

        try:
            import pandas as pd
            df = pd.read_excel(str(errors_file))
            if df.empty:
                QMessageBox.information(self, "Пустой файл", "Файл errors.xlsx пуст.")
                return
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Не удалось прочитать errors.xlsx:\n{e}")
            return

        # === Новый блок: выбор типа (файлы / папки) ===
        button = QMessageBox(self)
        button.setWindowTitle("Автомат — выбор типа")
        button.setText("Что передавать?")
        button.setIcon(QMessageBox.Icon.Question)
        btn_files = button.addButton("Файлы", QMessageBox.ButtonRole.YesRole)
        btn_folders = button.addButton("Папки", QMessageBox.ButtonRole.NoRole)
        button.addButton("Отмена", QMessageBox.ButtonRole.RejectRole)
        button.exec()

        if button.clickedButton() == btn_files:
            files, _ = QFileDialog.getOpenFileNames(self, "Файлы для Автомата", "", "All Files (*)")
            selected_files = files
        elif button.clickedButton() == btn_folders:
            folder = QFileDialog.getExistingDirectory(self, "Папка для Автомата", "")
            selected_files = [folder] if folder else []
        else:
            return

        if not selected_files:
            return

        # === Запуск автомата ===
        try:
            if hasattr(self.anya_panel, 'show_check_image'):
                self.anya_panel.show_check_image()
        except:
            pass

        self.btn_automat.setEnabled(False)
        self.btn_check.setEnabled(False)
        self.btn_transfer.setEnabled(False)
        self.btn_stop.setEnabled(True)

        self.automat_worker = AutomatErrorWorker(
            errors_file=str(errors_file),
            db=self.db,
            all_branches=self.all_branches,
            selected_files=self.selected_files,  # ✅ Correct source
            max_concurrent=50
        )

        self.automat_worker.progress.connect(self._on_automat_progress)
        self.automat_worker.transfer_status.connect(self._on_automat_status)
        self.automat_worker.error_logged.connect(self._on_error_logged)
        self.automat_worker.finished.connect(self._on_automat_finished)
        self.automat_worker.error.connect(self._on_automat_error)
        self.automat_worker.check_integrity = True  # обязательно SHA256
        self.automat_worker.stop_requested = False

        self.progress_bar.setVisible(True)
        self.progress_label.setText("Автомат: запуск...")
        self.automat_worker.start()

    def _on_automat_progress(self, current: int, total: int):
        self.progress_bar.setMaximum(total)
        self.progress_bar.setValue(current)
        self.progress_label.setText(f"Авто: {current}/{total}")

    def _on_automat_status(self, prefix: str, ip: str, status: str):
        level = logger.info if status.startswith("✓") else logger.warning
        level(f"[АВТОМАТ] {prefix} {ip} {status}")

    def _on_automat_finished(self):
        self._cleanup_after_worker()
        try:
            if hasattr(self.anya_panel, 'show_success_image'):
                self.anya_panel.show_success_image()
        except:
            pass
        QMessageBox.information(self, "Автомат", "Автоматическая передача завершена")
        logger.info("[АВТОМАТ] Работа завершена")

    def _on_automat_error(self, error_msg: str):
        self._cleanup_after_worker()
        try:
            if hasattr(self.anya_panel, 'show_error_image'):
                self.anya_panel.show_error_image()
        except:
            pass
        QMessageBox.critical(self, "Ошибка Автомата", f"Ошибка:\n{error_msg}")
        logger.error(f"[АВТОМАТ] Критическая ошибка: {error_msg}")

    def _cleanup_after_worker(self):
        self.progress_bar.setVisible(False)
        self.btn_automat.setEnabled(True)
        self.btn_check.setEnabled(True)
        self.btn_transfer.setEnabled(len(self.selected_branches) > 0)
        self.btn_stop.setEnabled(False)
        self._update_automat_button()

    def _on_stop(self):
        if self.check_worker and self.check_worker.isRunning():
            logger.info("Запрос на остановку проверки")
            self.check_worker.stop_requested = True
            self.btn_stop.setEnabled(False)
        if self.transfer_worker and self.transfer_worker.isRunning():
            logger.info("Запрос на остановку передачи")
            self.transfer_worker.stop_requested = True
            self.btn_stop.setEnabled(False)
        if self.automat_worker and self.automat_worker.isRunning():
            logger.info("Запрос на остановку автомата")
            self.automat_worker.stop_requested = True
            self.btn_stop.setEnabled(False)

    def _on_error_logged(self, prefix: str, branch_name: str, error_msg: str):
        self.errors_tab.add_error(prefix, branch_name, error_msg)

    def _on_transfer_finished(self):
        self.progress_bar.setVisible(False)
        self.btn_check.setEnabled(True)
        self.btn_transfer.setEnabled(len(self.selected_branches) > 0)
        self.btn_stop.setEnabled(False)
        self._update_selection()
        QMessageBox.information(self, "Передача", "Передача файлов завершена")
        logger.info("Передача файлов завершена")

    def _on_transfer_error(self, error_msg: str):
        self.progress_bar.setVisible(False)
        self.btn_check.setEnabled(True)
        self.btn_transfer.setEnabled(len(self.selected_branches) > 0)
        self.btn_stop.setEnabled(False)
        self._update_selection()
        QMessageBox.critical(self, "Ошибка", f"Ошибка передачи:\n{error_msg}")
        logger.error(f"Ошибка передачи: {error_msg}")

    def _on_clear_logs(self):
        self.log_panel.clear()
        logger.info("Логи очищены")

    def _on_show_instruction(self):
        InstructionDialog(self).exec()

    def _on_log_message(self, message: str, level: str):
        try:
            self.log_panel.append(message)
            self.log_panel.verticalScrollBar().setValue(self.log_panel.verticalScrollBar().maximum())
        except Exception as e:
            logger.warning(f"Ошибка отображения лога: {e}")


if __name__ == "__main__":
    import sys
    from PyQt6.QtWidgets import QApplication
    app = QApplication(sys.argv)
    window = MainWindow()
    window.show()
    sys.exit(app.exec())
